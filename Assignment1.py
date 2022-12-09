# I did the assignment by my phone, some cases without test. I did it as far as I can, with the limitations.

from openpyxl import workbook, load_workbook
from selenium import webdriver
browser = webdriver.Chrome()
# Excel file path in my device
wk = load_workbook("/storage/emulated/0/Download/Excel.xlsx")
#loop through sheet
for n, sheet in enumerate(wk.worksheets):
    sh = wk.active
    #loop through rows in specific column every active sheet
    for row in sh.iter_rows(min_row = 3, min_col = 3, max_row =12, max_col = 3):
        for cell in row:
            search_term = cell.value 
            
          #function for every cell search option
            def get_result(search_term):
                url = 'https://www.google.com'
                browser.get(url)
                search_box = browser.find_element_by_id('query'')
                search_box.send_keys(search_term)
                search_box.submit()
                longest_option = browser.find_element_by_xpath('result')
                shortest_option = browser.find_element_by_xpath('result')
                # following lines are data insert into sheet
                r = 3
                sh.cell(row = r, column = 4).value = longest_option
                sh.cell(row = r, column = 5).value = shortest_option
                r += 1
                
                return sh
                
            get_result(search_term) 
       
browser.close()
wk.save("/storage/emulated/0/Download/Excel.xlsx")